"""ADD-5 — Microsoft Project Online → Helm import wizard.

Project Online (PWA) end-of-sale 2025-10-01, retirement 2026-09-30.
Microsoft's only direct successor (Planner Premium) is not in GCC,
no committed parity date. ~200+ GCC PWA customers must migrate before
September 30, 2026. This module is the migration on-ramp.

Accepts a CSV or Excel export from a PWA Projects view and creates one
Helm Project per row, mapping standard PWA columns to Helm fields:

    PWA column                Helm field
    ----------                ----------
    Project Name / Name       Project.name
    Description               Project.description
    Start                     Project.started_at
    Finish                    Project.target_end_at
    % Complete                — informational only (we recompute from tasks)
    Owner                     — informational only (assignment is manual after import)
    Notes                     Created as a single ProjectNote on the imported project

Each row goes through ``services.create_project()`` so the import is
audit-logged, ACL-respected, and notification-fired exactly like a
manual creation. Rows that fail validation surface in a structured
report and don't block other rows.

"Trial" mode runs the parse + dry-run pass without writing to the DB
so the user sees the column mapping + row-count report before
committing. "Commit" mode actually creates the projects.

What this DOESN'T import:
- Resources / assignments — too varied across PWA setups; manual.
- Custom fields — same.
- Sub-tasks / WBS hierarchy — Helm tasks are flat in v1.
- Linked SharePoint sites — keep them in SharePoint.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Optional


# Standard PWA column aliases — case-insensitive match by lowered key.
_NAME_KEYS = ('project name', 'name', 'project')
_DESC_KEYS = ('description', 'project description', 'desc')
_START_KEYS = ('start', 'start date', 'project start')
_FINISH_KEYS = ('finish', 'finish date', 'end date', 'target finish')
_NOTES_KEYS = ('notes', 'comments', 'project notes')
_OWNER_KEYS = ('owner', 'project owner', 'manager')


@dataclass
class ImportRow:
    """One parsed row from the PWA export."""
    raw: dict
    name: str
    description: str = ''
    started_at: Optional[date] = None
    target_end_at: Optional[date] = None
    notes_body: str = ''
    owner_label: str = ''
    errors: list[str] = None  # type: ignore[assignment]

    def __post_init__(self):
        if self.errors is None:
            self.errors = []

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass
class ImportReport:
    """Summary of a parse pass."""
    rows: list[ImportRow]
    column_mapping: dict[str, str]   # source column → Helm field
    unmapped_columns: list[str]
    parse_errors: list[str]

    @property
    def valid_count(self) -> int:
        return sum(1 for r in self.rows if r.is_valid)

    @property
    def error_count(self) -> int:
        return sum(1 for r in self.rows if not r.is_valid)


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def _normalize_key(s: str) -> str:
    return (s or '').strip().lower()


def _resolve_field(row: dict, candidates: tuple[str, ...]) -> Optional[str]:
    """Look up a value by trying each candidate column name (case-insensitive)."""
    lower_row = {_normalize_key(k): v for k, v in row.items()}
    for candidate in candidates:
        if candidate in lower_row:
            v = lower_row[candidate]
            if v is None:
                return None
            return str(v).strip() if not isinstance(v, str) else v.strip()
    return None


def _parse_date(value) -> Optional[date]:
    """Best-effort date parsing. PWA exports use varied formats."""
    if value in (None, ''):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_csv(content: bytes) -> ImportReport:
    """Parse a UTF-8 CSV (or BOM-prefixed) PWA export."""
    text = content.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    return _parse_dictreader(reader)


def parse_xlsx(content: bytes) -> ImportReport:
    """Parse an Excel (.xlsx) PWA export. Reads the first sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ImportReport(
            rows=[], column_mapping={}, unmapped_columns=[],
            parse_errors=['openpyxl not installed — cannot parse Excel files.'],
        )
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else '' for c in next(rows_iter)]
    except StopIteration:
        return ImportReport(
            rows=[], column_mapping={}, unmapped_columns=[],
            parse_errors=['Empty workbook.'],
        )
    dict_rows = []
    for row in rows_iter:
        if all(c is None or c == '' for c in row):
            continue  # blank row
        dict_rows.append(dict(zip(header, row)))
    # Use a dict-based reader-shaped iterable for shared logic.
    return _parse_dict_rows(header, dict_rows)


def _parse_dictreader(reader: csv.DictReader) -> ImportReport:
    rows = list(reader)
    header = list(reader.fieldnames or [])
    return _parse_dict_rows(header, rows)


def _parse_dict_rows(header: list[str], dict_rows: list[dict]) -> ImportReport:
    # Compute column mapping for the report.
    mapping: dict[str, str] = {}
    for col in header:
        norm = _normalize_key(col)
        if norm in _NAME_KEYS:
            mapping[col] = 'name'
        elif norm in _DESC_KEYS:
            mapping[col] = 'description'
        elif norm in _START_KEYS:
            mapping[col] = 'started_at'
        elif norm in _FINISH_KEYS:
            mapping[col] = 'target_end_at'
        elif norm in _NOTES_KEYS:
            mapping[col] = 'notes (created as ProjectNote)'
        elif norm in _OWNER_KEYS:
            mapping[col] = '— (owner; manual assignment after import)'
    unmapped = [c for c in header if c not in mapping and c]

    parsed_rows = []
    for raw in dict_rows:
        name = _resolve_field(raw, _NAME_KEYS)
        row = ImportRow(
            raw=raw,
            name=name or '',
            description=_resolve_field(raw, _DESC_KEYS) or '',
            started_at=_parse_date(_resolve_field(raw, _START_KEYS)),
            target_end_at=_parse_date(_resolve_field(raw, _FINISH_KEYS)),
            notes_body=_resolve_field(raw, _NOTES_KEYS) or '',
            owner_label=_resolve_field(raw, _OWNER_KEYS) or '',
        )
        if not row.name:
            row.errors.append('Missing project name (required).')
        parsed_rows.append(row)

    return ImportReport(
        rows=parsed_rows,
        column_mapping=mapping,
        unmapped_columns=unmapped,
        parse_errors=[],
    )


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------
def commit_import(report: ImportReport, *, user) -> dict:
    """Materialize valid rows as Helm projects. Returns a result dict.

    Each project is created via ``services.create_project()`` so the
    audit log, ACL, and notification pipeline all fire as if the user
    created each project by hand.
    """
    from tasks.services import add_project_note, create_project

    created = []
    skipped = []
    failed = []

    for row in report.rows:
        if not row.is_valid:
            skipped.append({'name': row.name, 'reasons': row.errors})
            continue
        try:
            project = create_project(
                name=row.name, user=user,
                description=row.description,
                started_at=row.started_at,
                target_end_at=row.target_end_at,
            )
            if row.notes_body:
                add_project_note(
                    project=project, user=user, content=row.notes_body,
                )
            created.append({
                'slug': project.slug, 'name': project.name,
                'public_id': str(project.public_id),
            })
        except Exception as e:
            failed.append({'name': row.name, 'error': str(e)})

    return {
        'created_count': len(created),
        'skipped_count': len(skipped),
        'failed_count': len(failed),
        'created': created,
        'skipped': skipped,
        'failed': failed,
    }
